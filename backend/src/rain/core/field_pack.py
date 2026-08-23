"""Best-effort type inference for a "field pack" import -- a spreadsheet
whose header row names a set of custom fields to define in bulk, with a
few rows of real/representative data underneath each column used only to
*guess* a field_type. Nothing here persists a cell value anywhere --
rain.modules.tickets.importer/rain.modules.assets.importer (that import
ticket/asset *rows*) are the ones that do that; this only ever produces
rain.db.tenant_models.CustomField definitions, reused by both modules'
"Import field pack" screens so the guessing logic doesn't drift between
the two.

Guessing is intentionally conservative -- everything it produces is an
editable suggestion on a preview screen (key/label/type/select options
are all just pre-filled form fields), never applied unseen, so a wrong
guess costs a click to fix rather than silently defining the wrong kind
of field."""
from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from io import BytesIO

from openpyxl import load_workbook

_BOOL_VALUES = {"true", "false", "yes", "no", "y", "n", "1", "0"}
# ISO (2024-01-31) or US-slashed (1/31/2024, 01/31/2024) -- the two
# formats a spreadsheet's own date-formatted cell round-trips to text as
# most often; anything else is left as "text" rather than guessed wrong.
_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$|^\d{1,2}/\d{1,2}/\d{4}$")
_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
_URL_RE = re.compile(r"^https?://", re.IGNORECASE)
_KEY_STRIP_RE = re.compile(r"[^a-z0-9_]+")

#: Sample rows read below the header -- enough to make a select-vs-text
#: call (needs to see repeats) without reading an entire large sheet just
#: to guess types from it.
_SAMPLE_ROWS = 20


@dataclass
class GuessedField:
    header: str
    field_key: str
    label: str
    field_type: str
    select_options: list[str] | None
    samples: list[str]


def sniff_columns(raw: bytes, fmt: str) -> list[GuessedField]:
    """fmt: "xlsx" | "csv". Reads the header row as column names and up to
    _SAMPLE_ROWS data rows underneath as samples -- a header-only file (or
    one with no data rows at all) still returns one GuessedField per
    header, just with field_type="text" (the same default the manual "New
    custom field" form starts on) and an empty samples list."""
    if fmt == "xlsx":
        headers, rows = _read_xlsx(raw)
    else:
        headers, rows = _read_csv(raw)

    fields = []
    for i, header in enumerate(headers):
        if header is None or not str(header).strip():
            continue
        samples = [
            str(row[i]).strip()
            for row in rows
            if i < len(row) and row[i] is not None and str(row[i]).strip()
        ]
        fields.append(_guess_field(str(header).strip(), samples))
    return fields


def _read_xlsx(raw: bytes) -> tuple[list, list[list]]:
    wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = list(next(rows_iter, ()))
    rows = []
    for row in rows_iter:
        rows.append(list(row))
        if len(rows) >= _SAMPLE_ROWS:
            break
    return headers, rows


def _read_csv(raw: bytes) -> tuple[list, list[list]]:
    text = raw.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    headers = next(reader, [])
    rows = []
    for row in reader:
        rows.append(row)
        if len(rows) >= _SAMPLE_ROWS:
            break
    return headers, rows


def slugify_key(header: str) -> str:
    """Same shape the manual "New custom field" form's field_key
    `pattern="[a-z][a-z0-9_]*"` requires -- a header that doesn't already
    look like that ("Warranty Expiry", "2024 Budget") gets lowercased,
    non-matching runs collapsed to "_", and, if it still doesn't start
    with a letter, prefixed with "f_" rather than rejected outright."""
    key = _KEY_STRIP_RE.sub("_", header.strip().lower()).strip("_")
    if not key or not key[0].isalpha():
        key = f"f_{key}" if key else "field"
    return key[:63]


def _is_number(value: str) -> bool:
    try:
        float(value)
        return True
    except ValueError:
        return False


def _guess_field(header: str, samples: list[str]) -> GuessedField:
    field_key = slugify_key(header)
    if not samples:
        return GuessedField(header, field_key, header, "text", None, samples)

    lowered = [s.lower() for s in samples]
    distinct = sorted(set(samples))

    if all(v in _BOOL_VALUES for v in lowered):
        field_type, options = "boolean", None
    elif all(_URL_RE.match(v) for v in samples):
        field_type, options = "url", None
    elif all(_EMAIL_RE.match(v) for v in samples):
        field_type, options = "email", None
    elif all(_DATE_RE.match(v) for v in samples):
        field_type, options = "date", None
    elif all(_is_number(v) for v in samples):
        field_type, options = "number", None
    elif len(samples) >= 4 and 1 < len(distinct) <= 6 and len(distinct) < len(samples):
        # A handful of values repeating across enough rows to look like a
        # fixed vocabulary rather than free text -- not applied to a
        # too-small sample, where every value trivially looks "distinct
        # enough" to pass this on its own.
        field_type, options = "select", distinct
    else:
        field_type, options = "text", None

    return GuessedField(header, field_key, header, field_type, options, samples)
